#!/usr/bin/env python3
"""
导入微信文章 Excel 到数据库并更新网站

使用方法：
    python3 import_excel.py /Users/yuming.chen/Desktop/wechat_articles_20260618_143000.xlsx
    python3 import_excel.py  # 自动查找最新 Excel
"""

import sqlite3
import json
import os
import sys
import subprocess
from pathlib import Path
from datetime import datetime

DB_PATH = Path.home() / ".openclaw/cosmetic_articles.db"
PROJECT_DIR = Path.home() / ".openclaw/workspace/cosmetic-deploy"

def ensure_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS articles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            url TEXT UNIQUE NOT NULL,
            source TEXT NOT NULL,
            publish_date TEXT,
            summary TEXT,
            keywords TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    conn.close()

def find_latest_excel():
    desktop = Path.home() / "Desktop"
    files = list(desktop.glob("wechat_articles_*.xlsx"))
    return sorted(files, key=lambda p: p.stat().st_mtime, reverse=True)[0] if files else None

def import_excel(excel_path):
    try:
        import openpyxl
    except ImportError:
        print("❌ 需要安装 openpyxl: pip install openpyxl")
        return 0
    
    conn = sqlite3.connect(DB_PATH)
    added = 0
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # 读取表头
    headers = [cell.value for cell in ws[1]]
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        
        account = data.get('公众号', '')
        link = data.get('链接', '')
        keyword = data.get('关键词', '-')
        
        if not link or not link.startswith('http'):
            continue
        
        title = f"{account} 文章"
        if keyword and keyword != '-':
            title += f" ({keyword})"
        
        try:
            conn.execute("""
                INSERT INTO articles (title, url, source, publish_date, keywords, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (title, link, account, datetime.now().strftime("%Y-%m-%d"),
                  json.dumps([keyword] if keyword != '-' else []),
                  datetime.now().isoformat()))
            added += 1
        except sqlite3.IntegrityError:
            pass
    
    conn.commit()
    conn.close()
    return added

def export_and_push():
    os.chdir(PROJECT_DIR)
    
    # 导出数据
    subprocess.run(["python3", "export_data.py"], check=True)
    
    # 推送 GitHub
    subprocess.run(["git", "add", "data.json"], check=True)
    subprocess.run(["git", "commit", "-m", f"更新: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"], check=True)
    subprocess.run(["git", "push", "origin", "main"], check=True)

def main():
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    else:
        excel_path = find_latest_excel()
        if not excel_path:
            print("❌ 未找到 wechat_articles_*.xlsx 文件")
            print("请先运行 wechat-article-collector 采集文章")
            sys.exit(1)
        print(f"📁 自动找到: {excel_path}")
    
    if not os.path.exists(excel_path):
        print(f"❌ 文件不存在: {excel_path}")
        sys.exit(1)
    
    print("📥 导入 Excel...")
    ensure_db()
    added = import_excel(excel_path)
    print(f"✅ 新增 {added} 篇")
    
    print("\n🚀 更新网站...")
    export_and_push()
    
    print("\n✅ 完成！")
    print("https://benavidesbraian710-hub.github.io/cosmetic-articles-search/")

if __name__ == "__main__":
    main()
